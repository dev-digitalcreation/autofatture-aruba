# -*- coding: utf-8 -*-
"""
Lettura/scrittura della configurazione (azienda + fornitori).

In esecuzione da sorgente: i file stanno nella cartella dell'app (config/).
In esecuzione da .exe (frozen): i file stanno in una cartella scrivibile
dell'utente (%APPDATA%\\Reversa), perche' dentro il pacchetto sono
in sola lettura. Se mancano vengono creati dai default incorporati.
"""
import os
import re
import sys
import json
import shutil

# --------------------------------------------------------------------------- #
# Default incorporati (usati alla prima esecuzione / installazione)
# --------------------------------------------------------------------------- #
# NB: default GENERICI (nessun dato personale, il repo e' pubblico).
# Ogni utente inserisce i propri dati azienda/IBAN in Impostazioni alla prima apertura.
# Trasmittente e canale SdI NON sono precompilati: l'utente sceglie il proprio
# intermediario dai preset qui sotto (Impostazioni -> Azienda -> "preset routing").

# Preset di routing/intermediario SdI (trasmittente + canale destinatario).
# Sono identificativi PUBBLICI dell'intermediario, uguali per tutti i suoi utenti.
# Estendibile: aggiungere qui altri intermediari in futuro.
ROUTING_PRESETS = {
    "Aruba": {
        "trasmittente_id_paese": "IT",
        "trasmittente_id_codice": "01879020517",
        "codice_destinatario": "KRRH6B9",
    },
}

DEFAULT_CONFIG = {
    "azienda": {
        "denominazione": "", "id_paese": "IT", "piva": "", "codice_fiscale": "",
        "indirizzo": "", "numero_civico": "", "cap": "", "comune": "", "provincia": "", "nazione": "IT",
    },
    "trasmittente": {"id_paese": "IT", "id_codice": ""},
    "codice_destinatario": "0000000",
    "soggetto_emittente": "CC",
    "importo_totale_documento": True,
    "numerazione": {"pattern": "{n}/{yy}"},
    "progressivo_start": 1,
    "filename_prefix": None,
    "ricorda_ultimo_numero": True,
    "aliquota_default": "22.00",
    "pagamento": None,          # opzionale; l'IBAN lo inserisce l'utente
    "github_token": "",
}

DEFAULT_FORNITORI = {
    "anthropic": {"denominazione": "Anthropic, PBC", "id_paese": "US", "id_codice": "OO99999999999",
                  "indirizzo": "548 Market Street PMB 90375", "cap": "00000", "comune": "San Francisco",
                  "nazione": "US", "tipo_documento": "TD17"},
    "siteground": {"denominazione": "SiteGround Spain S.L.", "id_paese": "ES", "id_codice": "B87194171",
                   "indirizzo": "28004 Calle Prim", "numero_civico": "19", "cap": "00000",
                   "comune": "Madrid", "nazione": "ES", "tipo_documento": "TD17"},
    "vercel": {"denominazione": "Vercel Inc.", "id_paese": "US", "id_codice": "OO99999999999",
               "indirizzo": "440 N Barranca Ave #4133", "cap": "00000", "comune": "Covina",
               "nazione": "US", "tipo_documento": "TD17"},
    "openai": {"denominazione": "OpenAI, LLC", "id_paese": "US", "id_codice": "OO99999999999",
               "indirizzo": "3180 18th Street", "cap": "00000", "comune": "San Francisco",
               "nazione": "US", "tipo_documento": "TD17"},
    "google": {"denominazione": "Google Cloud EMEA Limited", "id_paese": "IE", "id_codice": "IE3668997OH",
               "indirizzo": "70 Sir John Rogerson's Quay", "cap": "00000", "comune": "Dublin",
               "nazione": "IE", "tipo_documento": "TD17"},
    "microsoft": {"denominazione": "Microsoft Ireland Operations Ltd", "id_paese": "IE",
                  "id_codice": "IE8256796U", "indirizzo": "One Microsoft Place", "cap": "00000",
                  "comune": "Dublin", "nazione": "IE", "tipo_documento": "TD17"},
    "amazon web services": {"denominazione": "Amazon Web Services EMEA SARL", "id_paese": "LU",
                            "id_codice": "LU26888617", "indirizzo": "38 avenue John F. Kennedy",
                            "cap": "00000", "comune": "Luxembourg", "nazione": "LU",
                            "tipo_documento": "TD17"},
}


# --------------------------------------------------------------------------- #
# Percorsi (scrivibili anche quando l'app e' un .exe)
# --------------------------------------------------------------------------- #
def _base_dir() -> str:
    if getattr(sys, "frozen", False):          # in esecuzione come .exe
        root = os.environ.get("APPDATA") or os.path.expanduser("~")
        base = os.path.join(root, "Reversa")
        # Migrazione una-tantum dal vecchio nome (AutofattureAruba -> Reversa):
        # se la cartella nuova non esiste ancora ma c'e' quella vecchia, copia i dati.
        legacy = os.path.join(root, "AutofattureAruba")
        if not os.path.exists(base) and os.path.isdir(legacy):
            try:
                shutil.copytree(legacy, base)
            except Exception:
                pass
    else:                                       # in esecuzione da sorgente
        base = os.path.dirname(os.path.abspath(__file__))
    os.makedirs(os.path.join(base, "config"), exist_ok=True)
    os.makedirs(os.path.join(base, "output_xml"), exist_ok=True)
    return base


BASE = _base_dir()
CONFIG_PATH = os.path.join(BASE, "config", "azienda.json")
FORNITORI_PATH = os.path.join(BASE, "config", "fornitori.json")
OUTPUT_DIR = os.path.join(BASE, "output_xml")


def carica_config() -> dict:
    if not os.path.exists(CONFIG_PATH):
        salva_config(DEFAULT_CONFIG)
        return dict(DEFAULT_CONFIG)
    try:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return dict(DEFAULT_CONFIG)


def salva_config(cfg: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def carica_fornitori() -> dict:
    if not os.path.exists(FORNITORI_PATH):
        salva_fornitori(DEFAULT_FORNITORI)
        return dict(DEFAULT_FORNITORI)
    try:
        with open(FORNITORI_PATH, encoding="utf-8") as f:
            return json.load(f).get("fornitori", DEFAULT_FORNITORI)
    except Exception:
        return dict(DEFAULT_FORNITORI)


def salva_fornitori(forn: dict):
    with open(FORNITORI_PATH, "w", encoding="utf-8") as f:
        json.dump({"_nota": "Fornitori riconosciuti automaticamente dal testo del PDF.",
                   "fornitori": forn}, f, ensure_ascii=False, indent=2)


# --------------------------------------------------------------------------- #
# Import/Export fornitori (CSV / Excel .xlsx)
# --------------------------------------------------------------------------- #
# Colonne canoniche del foglio fornitori (la prima è la chiave).
FORNITORI_COLS = ["chiave", "denominazione", "id_paese", "id_codice", "indirizzo",
                  "numero_civico", "cap", "comune", "nazione", "tipo_documento"]


def _norm_h(h) -> str:
    """Normalizza un'intestazione: solo lettere/cifre minuscole (via spazi/punti/underscore)."""
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


# Alias di intestazione -> colonna canonica (tollerante a maiuscole/spazi/sinonimi).
_ALIAS_F = {_norm_h(c): c for c in FORNITORI_COLS}
# NB: "Nome"/"Cognome" NON mappano su denominazione: nell'export Aruba sono
# colonne distinte da "Denominazione" (nome/cognome della persona fisica).
_ALIAS_F.update({
    "paese": "id_paese",
    "piva": "id_codice", "partitaiva": "id_codice", "vat": "id_codice", "idvat": "id_codice",
    "ncivico": "numero_civico", "civico": "numero_civico",
    "tipodoc": "tipo_documento", "td": "tipo_documento", "tipo": "tipo_documento",
    "denom": "denominazione", "fornitore": "denominazione", "ragionesociale": "denominazione",
})


def fornitori_a_righe(forn: dict) -> list:
    """Dict fornitori -> lista di righe [{colonna: valore}], una per fornitore (ordinate)."""
    righe = []
    for chiave in sorted(forn):
        ana = forn.get(chiave) or {}
        r = {"chiave": chiave}
        for c in FORNITORI_COLS[1:]:
            r[c] = str(ana.get(c, "") or "")
        righe.append(r)
    return righe


def _righe_a_fornitori(righe) -> dict:
    """Righe (dict per colonna) -> dict fornitori. Tollerante su intestazioni
    (maiuscole/spazi); se manca la chiave la ricava dalla denominazione."""
    out = {}
    for raw in righe:
        r = {}
        for k, v in raw.items():
            col = _ALIAS_F.get(_norm_h(k))
            if not col:
                continue
            v = "" if v is None else v
            # Più intestazioni possono mappare sulla stessa colonna: tieni il
            # primo valore non vuoto, non farlo sovrascrivere da uno vuoto.
            if col not in r or (not str(r[col]).strip() and str(v).strip()):
                r[col] = v
        chiave = str(r.get("chiave", "")).strip().lower()
        denom = str(r.get("denominazione", "")).strip()
        if not chiave:
            chiave = denom.lower()
        if not chiave and not denom:
            continue  # riga vuota, salta
        ana = {}
        for c in FORNITORI_COLS[1:]:
            val = str(r.get(c, "")).strip()
            if val:
                ana[c] = val
        ana.setdefault("tipo_documento", "TD17")
        ana.setdefault("cap", "00000")
        out[chiave] = ana
    return out


def esporta_fornitori(forn: dict, path: str) -> str:
    """Esporta i fornitori in .csv o .xlsx (in base al suffisso di `path`). Ritorna il path."""
    righe = fornitori_a_righe(forn)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fornitori"
        ws.append(FORNITORI_COLS)
        for r in righe:
            ws.append([r.get(c, "") for c in FORNITORI_COLS])
        wb.save(path)
    else:  # CSV: ';' + BOM UTF-8 -> si apre correttamente in Excel italiano
        import csv
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=FORNITORI_COLS, delimiter=";")
            w.writeheader()
            for r in righe:
                w.writerow(r)
    return path


def importa_fornitori(path: str) -> dict:
    """Legge fornitori da .csv o .xlsx. Ritorna {chiave: anagrafica}."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {}
        header = [str(h or "").strip() for h in rows[0]]
        righe = [dict(zip(header, ["" if c is None else c for c in r])) for r in rows[1:]]
    else:
        import csv
        with open(path, encoding="utf-8-sig", newline="") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                delim = csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
            except Exception:
                delim = ";"
            righe = list(csv.DictReader(f, delimiter=delim))
    return _righe_a_fornitori(righe)
